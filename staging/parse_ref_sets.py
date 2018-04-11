'''parse_ref_sets -- parse CMS reference files

Usage:

    >>> argv = 'python parse_ref_sets ref1.xls'.split()

    >>> io = MockIO()
    >>> main(argv, io.cwd, io.load_workbook)

The output has a data file and a .ctl file for each table:

    >>> print io.fs['./ref_icd9.csv'].strip()
    ICD9:250,Diabetes

    >>> print(io.fs['./ref_icd9.ctl'])
    ... #doctest: +NORMALIZE_WHITESPACE
    load data append into table ref_icd9
    fields terminated by "," optionally enclosed by '"'
    trailing nullcols
    (code, description )

plus SQL scripts to create and drop all tables:

    >>> print(io.fs['./oracle_create_ref.sql'])
    ... #doctest: +NORMALIZE_WHITESPACE
    create table ref_icd9 (
      code VARCHAR2(16),
      description VARCHAR2(16)
    );

    >>> print(io.fs['./oracle_drop_ref.sql'])
    drop table ref_icd9;

and a shell script to run all the .ctl files:

    >>> print(io.fs['./sqlldr_all_ref.sh'])
    ... #doctest: +NORMALIZE_WHITESPACE
    sqlldr $SQLLDR_USER/$SQLLDR_PASSWORD@$ORACLE_SID control=./ref_icd9.ctl
           direct=true rows=1000000
           log=ref_icd9.csv.log bad=ref_icd9.csv.bad data=ref_icd9.csv
    <BLANKLINE>
    # The above should have loaded (not specified in the .fts) rows.

    >>> len(io.fs.keys())
    5

'''

from collections import OrderedDict
from datetime import datetime as datetime_T
from io import BytesIO, StringIO
from numbers import Number
import logging

try:
    from typing import (
        Any, Dict, Callable, IO, List,
        Optional as Opt, Tuple, Union, cast
    )
    from pathlib import Path as Path_T  # type: ignore
    Dict, Callable, IO, List, Opt, Tuple, Path_T
    Value = Union[str, Number, datetime_T, None]
    from parse_fts import Field0
    Workbook = Any  # kludge: no openpyxl stubs
    Cell = Any      # kludge
except ImportError:
    cast = lambda t, x: x  # type: ignore # noqa
    Field0 = None          # type: ignore

from unicodecsv import writer as csv_writer  # type: ignore

from parse_fts import SQLLoader, Field, LoadScript, FileInfo

log = logging.getLogger(__name__)


def main(argv, cwd, load_workbook):
    # type: (List[str], Path_T, Callable[[str], Workbook]) -> None
    load_script_cmds = []
    sql_data = []
    tables = []

    wb = load_workbook(argv[1])
    out = Output(cwd)
    for sheet_name in wb.get_sheet_names():
        sh = RefSheet(wb.get_sheet_by_name(sheet_name).rows, sheet_name)
        log.info('Processing %s', sh.table_name)
        fields, data_rows = sh.data()

        tables.append(sh.table_name)
        sql_data.append(Field.create_table(sh.table_name, fields))

        out.save_data(sh.table_name, data_rows)
        cmd = out.save_ctl(sh.table_name, fields)
        load_script_cmds.append(cmd)

    out.write_scripts(load_script_cmds, sql_data, tables)


class Output(object):
    sqlldr_script = 'sqlldr_all_ref.sh'
    sql_create = 'oracle_create_ref.sql'
    sql_drop = 'oracle_drop_ref.sql'

    def __init__(self, dest):
        # type: (Path_T) -> None
        self.__dest = dest

    @classmethod
    def csv_file_name(cls, table_name):
        # type: (str) -> str
        return table_name + '.csv'

    def save_data(self, table_name, data_rows):
        # type: (str, List[List[Value]]) -> None
        with (self.__dest / self.csv_file_name(table_name)).open('wb') as fout:
            w = csv_writer(fout)
            for row in data_rows:
                w.writerow(row)

    def save_ctl(self, table_name, fields):
        # type: (str, List[Field0]) -> str
        ctl_file = self.__dest / (table_name + '.ctl')
        csv_fn = self.csv_file_name(table_name)
        SQLLoader.save(ctl_file, table_name, fields)
        return LoadScript.sqlldr_cmd(ctl_file, csv_fn,
                                     FileInfo(csv_fn, None, None))

    def write_scripts(self, load_script_cmds, sql_data, tables):
        # type: (List[str], List[str], List[str]) -> None
        with (self.__dest / self.sqlldr_script).open('wb') as fout:
            fout.write('\n'.join(load_script_cmds))
        with (self.__dest / self.sql_create).open('wb') as fout:
            fout.write('\n\n'.join(sql_data))
        with (self.__dest / self.sql_drop).open('wb') as fout:
            fout.write('\n'.join(['drop table %s;' % t for t in tables]))


class RefSheet(object):
    def __init__(self, rows, sheet_name):
        # type: (List[List[Cell]], str) -> None
        self.rows = rows
        self.sheet_name = sheet_name
        self.table_name = 'ref_' + self.sheet_name.replace(' ', '_').lower()

    def data(self):
        # type: () -> Tuple[List[Field0], List[List[Value]]]
        header = {}  # type: Dict[str, Field]
        sheet_name = self.sheet_name

        def p2size(sz):
            # type: (int) -> int
            return 1 << (sz-1).bit_length()

        def update_type(head, new_type, idx):
            # type: (str, str, int) -> None
            old_type = header[head].dtype
            if old_type and old_type != new_type:
                raise RuntimeError('%s column changed to %s! %s, %d' %
                                   (old_type, new_type, sheet_name, idx))
            header[head] = header[head]._replace(dtype=new_type)

        out_rows = []  # type: List[List[Value]]
        for idx, row in enumerate(self.rows):
            # The first few lines may be description/title
            if not header:
                if None not in [c.value for c in row]:
                    header = self.header_fields(row)
            else:
                row_to_write = []  # type: List[Value]
                for cell, head in zip(row, header.keys()):
                    if isinstance(cell.value, datetime_T):
                        update_type(head, 'DATE', idx)
                        row_to_write.append(cell.value.strftime('%Y%m%d'))
                    elif isinstance(cell.value, Number):
                        update_type(head, 'NUM', idx)
                        row_to_write.append(cell.value)
                    elif cell.value:
                        update_type(head, 'CHAR', idx)
                        header[head] = header[head]._replace(
                            width=max(header[head].width,
                                      p2size(len(cell.value) +
                                             Field.MIN_VARCHAR2_LEN)))
                        row_to_write.append(cell.value.strip()
                                            if cell.value.strip()
                                            else None)
                    else:
                        row_to_write.append(None)
                if True in [c is not None for c in row_to_write]:
                    out_rows.append(row_to_write)

        fields = [cast(Field0, f) for f in header.values()]
        return fields, out_rows

    @classmethod
    def header_fields(cls, row):
        # type: (List[Any]) -> Dict[str, Field]
        def abbr_reserved(c):
            # type: (str) -> str
            ''' Replace Oracle reserved words.
            '''
            return 'levl' if c == 'level' else c

        return OrderedDict([
            (name, Field(name=name, dtype=None, label=None,
                         width=Field.MIN_VARCHAR2_LEN))
            for cell in row
            for name in [abbr_reserved(cell.value.replace(' ', '_').lower())]])


class MockIO(object):
    fs = {}  # type: Dict[str, bytes]

    def __init__(self, path=None):
        # type: (Opt[str]) -> None
        self.path = path or '.'

    def __str__(self):
        # type: () -> str
        return self.path

    @property
    def cwd(self):
        # type: () -> MockIO
        return MockIO('.')

    def __div__(self, other):
        # type: (str) -> Any
        return self.__class__(self.path + '/' + other)

    def open(self, mode='r'):
        # type: (str) -> IO[Any]
        if mode == 'r':
            return BytesIO(self.fs[self.path])
        elif mode == 'wb':
            def doneb(value):
                # type: (bytes) -> None
                self.fs[self.path] = value
            return MockFP(doneb)
        else:
            raise IOError(mode)

    def load_workbook(self, path):
        # type: (str) -> Workbook
        return self

    def get_sheet_names(self):
        # type: () -> List[str]
        return ['ICD9']

    def get_sheet_by_name(self, name):
        # type: (str) -> Any
        return self

    @property
    def rows(self):
        # type: () -> List[List[Any]]
        class Cell(object):
            def __init__(self, value):
                # type: (str) -> None
                self.value = value

        return [[Cell('Code'), Cell('Description')],
                [Cell('ICD9:250'), Cell('Diabetes')]]


class MockFP(BytesIO):
    def __init__(self, done):
        # type: (Callable[[bytes], None]) -> None
        BytesIO.__init__(self)
        self.done = done

    def close(self):
        # type: () -> None
        self.done(self.getvalue())


if __name__ == '__main__':
    def _tcb():
        # type: () -> None
        from sys import argv

        from openpyxl import load_workbook  # type: ignore
        from pathlib import Path

        logging.basicConfig(level=logging.INFO)

        main(argv, cwd=Path('.'), load_workbook=load_workbook)

    _tcb()
